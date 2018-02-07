import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
dict1 = []
dictResult = {}

def fetch(results):
    dict1 = []
    for item in results:
        if item.find_all("div",class_="sc-table-col sc-day-header sc-gray") == []:
            key = item.find_all("div",class_="sc-table-col sc-day-header sc-blue")[0].text
        else:
            key = item.find_all("div",class_="sc-table-col sc-day-header sc-gray")[0].text
        dictResult[key] = []
        items = item.find_all("div",class_="sc-table-row")
        for row in items:
            for i in row.text.split('\n'):
                if i != '' : dict1.append(i.encode())
            dictResult[key].append(dict1)
            dict1 = []
        dictResult[key].remove(dictResult[key][0])
    with open("F:\\scheduleKi2Nam1.txt",'wb') as f:
        for key in dictResult:
            if key[-2:] == 'Пн':
                f.write(b'----------------------------------------------------')
                f.write(b'\n')
            f.write(key.encode())
            f.write(b'\n')
            for row in dictResult[key]:
                f.write('\t\t\t'.encode())
                f.write(b"---".join(row))
                f.write(b'\n')

def writeExecl():
    row_index = 1
    for key in dictResult:
        ws.cell(row = row_index,column = 1,value = key)
        row_index += 2
        for row in dictResult[key]:
            for i,value in enumerate(row):
                ws.cell(row= row_index,column=i+1,value = value)
            row_index += 1
        row_index += 1
    wb.save("F:\\schedux.xlsx")
if __name__ == '__main__':           
    for i in range(2,18):
        try:
            html = 'https://mai.ru/education/schedule/detail.php?group=%D0%9C2%D0%9E-111%D0%91%D0%BA%D0%B8-17&week=' + str(i)
            #html = 'https://mai.ru/education/schedule/detail.php?group=%D0%9C2%D0%9E-111%D0%91%D0%BA%D0%B8-17&week=' + str(i)
            result = requests.get(html)
            soup = BeautifulSoup(result.text,'html.parser')
            results = soup.find_all("div",class_="sc-container")
            fetch(results)
        except:
            pass
        #writeExecl()
        